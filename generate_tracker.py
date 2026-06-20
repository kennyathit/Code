#!/usr/bin/env python3
"""Generate the TG (Thai Airways) Pilot Study Tracker spreadsheet.

Reads config.yaml and writes a multi-tab .xlsx that is meant to be opened /
imported as a Google Sheet. Members enter one row per practice attempt on the
"Scores" tab; everything else (scoreboard, weakest->strongest training
priority, daily schedule, gap analysis, submission status) computes itself.

Usage:
    python3 generate_tracker.py            # clean file, ready to use
    python3 generate_tracker.py --demo     # extra file seeded with sample data
"""

from __future__ import annotations

import argparse
import datetime as dt
import os

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Layout limits (kept generous for a small group) ───────────────────────────
MAX_MEMBERS = 20          # Config A2:A21  /  Scoreboard rows 2..21
MAX_SUBJECTS = 30         # Config C2:C31  /  Scoreboard cols B..AE (2..31)
MAX_BOOKS = 40            # Config F2:F41
SCORE_ROWS = 1000         # Scores rows 2..1001

MEMBER_FIRST, MEMBER_LAST = 2, 2 + MAX_MEMBERS - 1            # 2..21
GROUP_ROW = MEMBER_LAST + 1                                   # 22
SUBJ_FIRST_COL, SUBJ_LAST_COL = 2, 2 + MAX_SUBJECTS - 1       # B..AE
OVERALL_COL = SUBJ_LAST_COL + 1                               # AF
SCORE_LAST = SCORE_ROWS + 1                                   # 1001

SUBJ_FIRST_L = get_column_letter(SUBJ_FIRST_COL)             # B
SUBJ_LAST_L = get_column_letter(SUBJ_LAST_COL)               # AE
OVERALL_L = get_column_letter(OVERALL_COL)                   # AF

# Frequently used absolute ranges -------------------------------------------------
R_MEMBERS = f"Config!$A${MEMBER_FIRST}:$A${MEMBER_LAST}"
R_SUBJECTS = f"Config!$C${MEMBER_FIRST}:$C${SUBJ_FIRST_COL + MAX_SUBJECTS - 1}"  # C2:C31
R_TARGETS = f"Config!$D$2:$D${MAX_SUBJECTS + 1}"
R_BOOKS = f"Config!$F$2:$F${MAX_BOOKS + 1}"

SC_MEMBERS = f"Scoreboard!$A${MEMBER_FIRST}:$A${MEMBER_LAST}"
SC_SUBJECTS = f"Scoreboard!${SUBJ_FIRST_L}$1:${SUBJ_LAST_L}$1"
SC_MATRIX = f"Scoreboard!${SUBJ_FIRST_L}${MEMBER_FIRST}:${SUBJ_LAST_L}${MEMBER_LAST}"
SC_GROUP = f"Scoreboard!${SUBJ_FIRST_L}${GROUP_ROW}:${SUBJ_LAST_L}${GROUP_ROW}"

S_DATE = f"Scores!$A$2:$A${SCORE_LAST}"
S_MEMBER = f"Scores!$B$2:$B${SCORE_LAST}"
S_SUBJECT = f"Scores!$C$2:$C${SCORE_LAST}"
S_PCT = f"Scores!$G$2:$G${SCORE_LAST}"

PCT_FMT = "0%"
DATE_FMT = "yyyy-mm-dd"

# ── Styling helpers ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=14, color="2F5597")
NOTE_FONT = Font(italic=True, size=9, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER


def load_config(path: str) -> dict:
    if os.path.exists(path):
        with open(path) as fh:
            cfg = yaml.safe_load(fh) or {}
    else:
        cfg = {}
    cfg.setdefault("members", ["You", "Friend A", "Friend B"])
    cfg.setdefault(
        "subjects",
        ["Number Series", "Figure / Picture Series", "Shape Scanning",
         "Aircraft / Spatial Rotation"],
    )
    cfg.setdefault("default_target", 0.80)
    cfg.setdefault("group_target", 0.75)
    cfg.setdefault("books", [])
    cfg.setdefault("schedule_days", 7)
    cfg.setdefault("deadline_in_days", 7)
    return cfg


# ── Tab builders ─────────────────────────────────────────────────────────────
def build_config(ws, cfg):
    today = dt.date.today()
    ws["A1"] = "Members"
    ws["C1"] = "Subjects"
    ws["D1"] = "Target %"
    ws["F1"] = "Books"
    ws["H1"] = "Settings"
    ws["I1"] = "Value"
    for c in ("A1", "C1", "D1", "F1", "H1", "I1"):
        header(ws[c])

    for i, name in enumerate(cfg["members"][:MAX_MEMBERS]):
        ws.cell(row=2 + i, column=1, value=name)
    for i, subj in enumerate(cfg["subjects"][:MAX_SUBJECTS]):
        ws.cell(row=2 + i, column=3, value=subj)
        t = ws.cell(row=2 + i, column=4, value=cfg["default_target"])
        t.number_format = PCT_FMT
    for i, book in enumerate(cfg["books"][:MAX_BOOKS]):
        ws.cell(row=2 + i, column=6, value=book.get("name", ""))

    settings = [
        ("Submission Deadline", today + dt.timedelta(days=cfg["deadline_in_days"]), DATE_FMT),
        ("Group Target %", cfg["group_target"], PCT_FMT),
        ("Schedule Start (today)", today, DATE_FMT),
        ("Round Start", today, DATE_FMT),
    ]
    for i, (label, value, fmt) in enumerate(settings):
        ws.cell(row=2 + i, column=8, value=label).font = Font(bold=True)
        vc = ws.cell(row=2 + i, column=9, value=value)
        vc.number_format = fmt

    ws.cell(row=7, column=8,
            value="Edit these lists anytime — the dashboards read them live.").font = NOTE_FONT
    for col, width in {"A": 18, "B": 2, "C": 26, "D": 10, "E": 2,
                       "F": 26, "G": 2, "H": 22, "I": 14}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def build_scores(ws, cfg, demo):
    cols = ["Date", "Member", "Subject", "Book", "Correct", "Total", "%", "Notes"]
    for i, name in enumerate(cols, start=1):
        header(ws.cell(row=1, column=i, value=name))
    ws["J1"] = "← Add one row per practice attempt. Pick Member/Subject/Book from the dropdowns; type Correct & Total."
    ws["J1"].font = NOTE_FONT

    for r in range(2, SCORE_LAST + 1):
        ws.cell(row=r, column=1).number_format = DATE_FMT
        g = ws.cell(row=r, column=7, value=f"=IFERROR(E{r}/F{r},\"\")")
        g.number_format = PCT_FMT

    # Dropdowns (cross-sheet list validation).
    dv_member = DataValidation(type="list", formula1=f"={R_MEMBERS}", allow_blank=True)
    dv_subject = DataValidation(type="list", formula1=f"={R_SUBJECTS}", allow_blank=True)
    dv_book = DataValidation(type="list", formula1=f"={R_BOOKS}", allow_blank=True)
    ws.add_data_validation(dv_member)
    ws.add_data_validation(dv_subject)
    ws.add_data_validation(dv_book)
    dv_member.add(f"B2:B{SCORE_LAST}")
    dv_subject.add(f"C2:C{SCORE_LAST}")
    dv_book.add(f"D2:D{SCORE_LAST}")

    if demo:
        _seed_scores(ws, cfg)

    widths = {"A": 12, "B": 14, "C": 24, "D": 24, "E": 9, "F": 9, "G": 8, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _seed_scores(ws, cfg):
    """Sample data: 3 members, two rounds, with a clear weakest subject and a
    lagging member, so every dashboard visibly works."""
    today = dt.date.today()
    d1, d2 = today - dt.timedelta(days=7), today - dt.timedelta(days=1)
    members = cfg["members"][:3]
    subj = cfg["subjects"]
    book = cfg["books"][0]["name"] if cfg["books"] else ""
    # (member_idx, subject_idx, date, correct, total)
    rows = [
        (0, 0, d1, 14, 20), (0, 0, d2, 17, 20),   # You: Number Series improving
        (0, 1, d2, 16, 20), (0, 2, d2, 9, 20),    # weak at Shape Scanning
        (0, 3, d2, 15, 20),
        (1, 0, d1, 12, 20), (1, 0, d2, 13, 20),   # Friend A: lagging overall
        (1, 1, d2, 11, 20), (1, 2, d2, 8, 20),
        (1, 3, d2, 10, 20),
        (2, 0, d2, 18, 20), (2, 1, d2, 17, 20),   # Friend B: strong
        (2, 2, d2, 12, 20), (2, 3, d2, 16, 20),
    ]
    r = 2
    for mi, si, d, c, t in rows:
        if mi >= len(members) or si >= len(subj):
            continue
        ws.cell(row=r, column=1, value=d).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=members[mi])
        ws.cell(row=r, column=3, value=subj[si])
        ws.cell(row=r, column=4, value=book)
        ws.cell(row=r, column=5, value=c)
        ws.cell(row=r, column=6, value=t)
        r += 1


def build_books(ws, cfg):
    cols = ["Book", "Subject(s) it trains", "Where to get / link",
            "Sent to", "Date sent", "Status"]
    for i, name in enumerate(cols, start=1):
        header(ws.cell(row=1, column=i, value=name))
    for i, b in enumerate(cfg["books"][:MAX_BOOKS]):
        row = 2 + i
        ws.cell(row=row, column=1, value=b.get("name", ""))
        ws.cell(row=row, column=2, value=b.get("subjects", ""))
        ws.cell(row=row, column=3, value=b.get("where", ""))
        ws.cell(row=row, column=4, value=b.get("sent_to", ""))
        ws.cell(row=row, column=5, value=b.get("date_sent", ""))
        ws.cell(row=row, column=6, value=b.get("status", ""))
    for col, w in {"A": 28, "B": 30, "C": 26, "D": 18, "E": 12, "F": 16}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def build_scoreboard(ws):
    ws["A1"] = "Member \\ Subject"
    header(ws["A1"])
    for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
        idx = c - SUBJ_FIRST_COL  # 0-based subject index
        cell = ws.cell(row=1, column=c, value=f"=IFERROR(Config!C{2 + idx},\"\")")
        header(cell)
    header(ws.cell(row=1, column=OVERALL_COL, value="Overall %"))

    for r in range(MEMBER_FIRST, MEMBER_LAST + 1):
        idx = r - MEMBER_FIRST
        mcell = ws.cell(row=r, column=1, value=f"=IFERROR(Config!A{2 + idx},\"\")")
        mcell.font = Font(bold=True)
        for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
            cl = get_column_letter(c)
            f = (
                f"=IF(OR($A{r}=\"\",{cl}$1=\"\"),\"\","
                f"IFERROR(AVERAGEIFS({S_PCT},{S_MEMBER},$A{r},{S_SUBJECT},{cl}$1,"
                f"{S_DATE},MAXIFS({S_DATE},{S_MEMBER},$A{r},{S_SUBJECT},{cl}$1)),\"\"))"
            )
            cell = ws.cell(row=r, column=c, value=f)
            cell.number_format = PCT_FMT
        ov = ws.cell(row=r, column=OVERALL_COL,
                     value=f"=IF($A{r}=\"\",\"\",IFERROR(AVERAGE({SUBJ_FIRST_L}{r}:{SUBJ_LAST_L}{r}),\"\"))")
        ov.number_format = PCT_FMT
        ov.font = Font(bold=True)

    header(ws.cell(row=GROUP_ROW, column=1, value="Group Avg"))
    for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
        cl = get_column_letter(c)
        g = ws.cell(row=GROUP_ROW, column=c,
                    value=f"=IF({cl}$1=\"\",\"\",IFERROR(AVERAGE({cl}{MEMBER_FIRST}:{cl}{MEMBER_LAST}),\"\"))")
        g.number_format = PCT_FMT
        g.font = Font(bold=True)
        g.fill = SUBHEAD_FILL

    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num", mid_value=0.6, mid_color="FFEB84",
        end_type="num", end_value=1, end_color="63BE7B",
    )
    ws.conditional_formatting.add(
        f"{SUBJ_FIRST_L}{MEMBER_FIRST}:{SUBJ_LAST_L}{MEMBER_LAST}", rule)

    ws.column_dimensions["A"].width = 16
    for c in range(SUBJ_FIRST_COL, OVERALL_COL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B2"


def build_training_priority(ws):
    ws["A1"] = "Training Priority — drill weakest first"
    ws["A1"].font = TITLE_FONT
    headers = ["Rank", "Subject", "Group Avg %", "Target %", "Gap to Target", "Weakest Member"]
    for i, name in enumerate(headers, start=1):
        header(ws.cell(row=3, column=i, value=name))

    # Hidden helper table (columns H..L) — one row per subject.
    last = MAX_SUBJECTS + 1  # rows 2..31
    ws["H1"] = "(helper — safe to ignore)"
    ws["H1"].font = NOTE_FONT
    for i in range(MAX_SUBJECTS):
        r = 2 + i
        col_arr = f"INDEX({SC_MATRIX},0,MATCH(H{r},{SC_SUBJECTS},0))"
        ws.cell(row=r, column=8, value=f"=IFERROR(Config!C{2 + i},\"\")")
        ws.cell(row=r, column=9,
                value=f"=IF(H{r}=\"\",\"\",IFERROR(INDEX({SC_GROUP},MATCH(H{r},{SC_SUBJECTS},0)),\"\"))")
        ws.cell(row=r, column=10,
                value=f"=IF(H{r}=\"\",\"\",IFERROR(INDEX({R_TARGETS},MATCH(H{r},{R_SUBJECTS},0)),\"\"))")
        ws.cell(row=r, column=11,
                value=f"=IF(OR(H{r}=\"\",I{r}=\"\",J{r}=\"\"),\"\",I{r}-J{r})")
        ws.cell(row=r, column=12,
                value=(f"=IF(OR(H{r}=\"\",I{r}=\"\"),\"\","
                       f"IFERROR(INDEX({SC_MEMBERS},MATCH(MIN({col_arr}),{col_arr},0)),\"\"))"))

    # Visible sorted view spills from B4 (sorted ascending by Group Avg).
    ws["B4"] = (
        f"=IFERROR(SORT(FILTER($H$2:$L${last},($H$2:$H${last}<>\"\")*($I$2:$I${last}<>\"\")),2,TRUE),\"\")"
    )
    for i in range(MAX_SUBJECTS):
        r = 4 + i
        ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",ROW()-3)").alignment = CENTER
        ws.cell(row=r, column=3).number_format = PCT_FMT
        ws.cell(row=r, column=4).number_format = PCT_FMT
        ws.cell(row=r, column=5).number_format = PCT_FMT

    for col in "HIJKL":
        ws.column_dimensions[col].hidden = True
    for col, w in {"A": 7, "B": 26, "C": 13, "D": 11, "E": 13, "F": 18}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def build_daily_schedule(ws, cfg):
    ws["A1"] = '=IF(C1="","Add scores to see today\'s drill","Today (" & TEXT(Config!$I$4,"ddd dd mmm yyyy") & ") — drill:")'
    ws["A1"].font = TITLE_FONT
    ws["C1"] = "=IFERROR('Training Priority'!$B$4,IFERROR(Config!$C$2,\"\"))"
    ws["C1"].font = Font(bold=True, size=14, color="C00000")

    for i, name in enumerate(["Date", "Subject (weakest first)", "Group Avg % now", "Manual override"], start=1):
        header(ws.cell(row=3, column=i, value=name))

    days = max(1, int(cfg["schedule_days"]))
    for i in range(days):
        r = 4 + i
        if i == 0:
            ws.cell(row=r, column=1, value="=Config!$I$4")
        else:
            ws.cell(row=r, column=1, value=f"=A{r - 1}+1")
        ws.cell(row=r, column=1).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=(
            f"=IF($D{r}<>\"\",$D{r},"
            f"IFERROR(INDEX('Training Priority'!$B$4:$B${3 + MAX_SUBJECTS},{i + 1}),"
            f"IFERROR(INDEX({R_SUBJECTS},{i + 1}),\"\")))"))
        ws.cell(row=r, column=3, value=(
            f"=IF($B{r}=\"\",\"\",IFERROR(INDEX({SC_GROUP},MATCH($B{r},{SC_SUBJECTS},0)),\"\"))"))
        ws.cell(row=r, column=3).number_format = PCT_FMT

    ws.cell(row=4 + days + 1, column=1,
            value="Type a subject in 'Manual override' to pin a day.").font = NOTE_FONT
    for col, w in {"A": 14, "B": 28, "C": 16, "D": 22}.items():
        ws.column_dimensions[col].width = w


def build_gap_analysis(ws):
    ws["A1"] = "Gap vs Group Avg  (green = ahead, red = behind)"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Member \\ Subject"
    header(ws["A3"])
    for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
        idx = c - SUBJ_FIRST_COL
        header(ws.cell(row=3, column=c, value=f"=IFERROR(Config!C{2 + idx},\"\")"))

    first, last = 4, 4 + MAX_MEMBERS - 1
    for r in range(first, last + 1):
        idx = r - first
        ws.cell(row=r, column=1, value=f"=IFERROR(Config!A{2 + idx},\"\")").font = Font(bold=True)
        for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
            cl = get_column_letter(c)
            f = (
                f"=IF(OR($A{r}=\"\",{cl}$3=\"\"),\"\",IFERROR("
                f"INDEX({SC_MATRIX},MATCH($A{r},{SC_MEMBERS},0),MATCH({cl}$3,{SC_SUBJECTS},0))"
                f"-INDEX({SC_GROUP},MATCH({cl}$3,{SC_SUBJECTS},0)),\"\"))"
            )
            cell = ws.cell(row=r, column=c, value=f)
            cell.number_format = "+0%;-0%;0%"

    rule = ColorScaleRule(
        start_type="num", start_value=-0.25, start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.25, end_color="63BE7B",
    )
    ws.conditional_formatting.add(
        f"{SUBJ_FIRST_L}{first}:{SUBJ_LAST_L}{last}", rule)

    ws.column_dimensions["A"].width = 16
    for c in range(SUBJ_FIRST_COL, SUBJ_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B4"


def build_submission_status(ws):
    ws["A1"] = "Submission Status (current round)"
    ws["A1"].font = TITLE_FONT
    ws["F1"] = "Round Start"
    ws["G1"] = "=Config!$I$5"
    ws["G1"].number_format = DATE_FMT
    ws["F2"] = "Deadline"
    ws["G2"] = "=Config!$I$2"
    ws["G2"].number_format = DATE_FMT
    for c in ("F1", "F2"):
        ws[c].font = Font(bold=True)

    for i, name in enumerate(["Member", "Last submission", "Attempts this round", "Status"], start=1):
        header(ws.cell(row=3, column=i, value=name))

    first, last = 4, 4 + MAX_MEMBERS - 1
    for r in range(first, last + 1):
        idx = r - first
        ws.cell(row=r, column=1, value=f"=IFERROR(Config!A{2 + idx},\"\")").font = Font(bold=True)
        ws.cell(row=r, column=2, value=(
            f"=IF($A{r}=\"\",\"\",IF(MAXIFS({S_DATE},{S_MEMBER},$A{r})=0,\"—\","
            f"MAXIFS({S_DATE},{S_MEMBER},$A{r})))"))
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=3, value=(
            f"=IF($A{r}=\"\",\"\",COUNTIFS({S_MEMBER},$A{r},{S_DATE},\">=\"&Config!$I$5))"))
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=4, value=(
            f"=IF($A{r}=\"\",\"\",IF(C{r}>0,\"submitted ✓\","
            f"IF(TODAY()>Config!$I$2,\"LATE — missing\",\"pending\")))"))

    for col, w in {"A": 16, "B": 16, "C": 18, "D": 18, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# ── Assemble ─────────────────────────────────────────────────────────────────
def build_workbook(cfg, demo=False) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_scores(wb.create_sheet("Scores"), cfg, demo)
    build_scoreboard(wb.create_sheet("Scoreboard"))
    build_training_priority(wb.create_sheet("Training Priority"))
    build_daily_schedule(wb.create_sheet("Daily Schedule"), cfg)
    build_gap_analysis(wb.create_sheet("Gap Analysis"))
    build_submission_status(wb.create_sheet("Submission Status"))
    build_books(wb.create_sheet("Books"), cfg)
    build_config(wb.create_sheet("Config"), cfg)
    return wb


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--demo", action="store_true",
                    help="also write a copy seeded with sample data")
    ap.add_argument("--config", default="config.yaml")
    args = ap.parse_args()

    cfg = load_config(args.config)

    clean = "TG_Pilot_Study_Tracker.xlsx"
    build_workbook(cfg, demo=False).save(clean)
    print(f"Wrote {clean}")

    if args.demo:
        demo_file = "TG_Pilot_Study_Tracker_DEMO.xlsx"
        build_workbook(cfg, demo=True).save(demo_file)
        print(f"Wrote {demo_file} (seeded with sample scores)")


if __name__ == "__main__":
    main()
